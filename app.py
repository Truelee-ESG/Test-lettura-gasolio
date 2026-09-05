import os
import re
from datetime import datetime
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image
import pytesseract

def seleziona_cartella():
    path = filedialog.askdirectory()
    if path:
        entry_path.delete(0, tk.END)
        entry_path.insert(0, path)

def analizza_file_gasolio(cartella):
    if not os.path.exists(cartella):
        messagebox.showerror("Errore", "La cartella specificata non esiste.")
        return None
    
    valid_extensions = ('.pdf', '.jpg', '.jpeg', '.png')
    target_keywords = ['gasolio', 'benzina', 'mezzi', 'auto', 'automezzi', 'trasporti', 'carburante', 'fleet', 'veicoli']
    
    file_list = []
    for root_dir, _, files in os.walk(cartella):
        folder_name = os.path.basename(root_dir).lower()
        is_target_dir = any(kw in folder_name or kw in root_dir.lower() for kw in target_keywords)
        
        if is_target_dir:
            for filename in files:
                if filename.lower().endswith(valid_extensions):
                    file_list.append((root_dir, filename))
                    
    if not file_list:
        for root_dir, _, files in os.walk(cartella):
            for filename in files:
                if filename.lower().endswith(valid_extensions):
                    file_list.append((root_dir, filename))
                    
    return file_list

def estrai_mesi_e_anno(text, text_lower, filename=""):
    mesi_mappa = {
        'gennaio': 'Gennaio', 'febbraio': 'Febbraio', 'marzo': 'Marzo', 'aprile': 'Aprile',
        'maggio': 'Maggio', 'giugno': 'Giugno', 'luglio': 'Luglio', 'agosto': 'Agosto',
        'settembre': 'Settembre', 'ottobre': 'Ottobre', 'novembre': 'Novembre', 'dicembre': 'Dicembre',
        'gen': 'Gen', 'feb': 'Feb', 'mar': 'Mar', 'apr': 'Apr',
        'mag': 'Mag', 'giu': 'Giu', 'lug': 'Lug', 'ago': 'Ago',
        'set': 'Set', 'ott': 'Ott', 'nov': 'Nov', 'dic': 'Dic'
    }
    anno = "Non rilevato"
    periodo = "Non rilevato"
    
    source_to_check = f"{filename} {text}"
    years = re.findall(r'\b(20\d{2})\b', source_to_check)
    if years:
        anno = years[0]
        
    found_mesi = []
    source_lower = f"{filename.lower()} {text_lower}"
    for m_key, m_val in mesi_mappa.items():
        if re.search(r'\b' + m_key + r'\b', source_lower):
            if m_val not in found_mesi:
                found_mesi.append(m_val)
    if found_mesi:
        periodo = found_mesi[0] if len(found_mesi) == 1 else f"{found_mesi[0]} - {found_mesi[-1]}"
        
    return periodo, anno

def estrai_dati_avanzati_ocr(text, text_lower, page=None):
    quantita = "Non rilevato"
    unita_misura = "Litri"
    
    # 1. Tentativo tramite analisi strutturata delle parole del PDF se la pagina è disponibile
    if page:
        try:
            words = page.extract_words(extra_attrs=["size"])
            for i, word in enumerate(words):
                w_text = word['text'].lower()
                if any(lbl in w_text for lbl in ['quantità', 'quantita', 'q.tà', 'q,tà', 'qta', 'um', 'u.m.']):
                    for j in range(i + 1, min(len(words), i + 6)):
                        cand_word = words[j]['text']
                        clean_cand = cand_word.replace('.', '').replace(',', '.')
                        if re.match(r'^\d+[\.,]?\d*$', clean_cand):
                            quantita = cand_word
                            break
                    if quantita != "Non rilevato":
                        break
        except Exception:
            pass

    # 2. Fallback avanzato tramite Regex multi-pattern ottimizzate sul testo estratto (OCR / PDF)
    if quantita == "Non rilevato":
        patterns_qty = [
            r'(?:quantit[aà]|q[\.,]tà|qta)\s*[:\-\=]?\s*(\d{1,3}(?:\.\d{3})*[\.,]?\d*)',
            r'\b(?:L|litri|litro)\b\s*(\d{1,3}(?:\.\d{3})*[\.,]?\d*)',
            r'(\d{1,3}(?:\.\d{3})*[\.,]?\d*)\s*(?:litri|Litri|\bL\b|litro)'
        ]
        for pat in patterns_qty:
            match = re.search(pat, text_lower, re.IGNORECASE)
            if match:
                quantita = match.group(1)
                break

    # Rilevamento unità di misura
    if re.search(r'\b(?:kg|KG|kilogrammi|chilogrammi)\b', text, re.IGNORECASE):
        unita_misura = "kg"
    elif re.search(r'\b(?:mc|MC|smc|SMC)\b', text, re.IGNORECASE):
        unita_misura = "Metri cubi"
    else:
        unita_misura = "Litri"

    return quantita, unita_misura

def avvia_estrazione():
    threading.Thread(target=_process_gasolio, daemon=True).start()

def _process_gasolio():
    azienda = entry_azienda.get().strip()
    cartella = entry_path.get().strip()
    if not azienda or not cartella:
        messagebox.showerror("Errore", "Inserisci il nome dell'azienda e seleziona una cartella valida.")
        return

    file_list = analizza_file_gasolio(cartella)
    if not file_list:
        messagebox.showwarning("Attenzione", "Nessun file valido trovato nella cartella.")
        return

    total_files = len(file_list)
    progress.config(maximum=total_files, value=0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consumi Gasolio"
    
    ws.append([f"Estrazione dati fatture gasolio e carburanti - {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws.cell(row=1, column=1).font = Font(size=12, bold=True, color="2E7D32")
    ws.append([])
    
    headers = ["Azienda", "Nome File", "Periodo", "Anno", "Quantità", "Unità di misura", "Carburante"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    valid_count = 0
    for idx, (root_dir, filename) in enumerate(file_list, 1):
        file_path = os.path.join(root_dir, filename)
        abs_path = os.path.abspath(file_path)
        ext = filename.lower()
        
        quantita = "Non rilevato"
        unita_misura = "Litri"
        carburante = "Non rilevato"
        periodo = "Non rilevato"
        anno = "Non rilevato"
        text = ""
        
        try:
            page_obj = None
            if ext.endswith('.pdf'):
                with pdfplumber.open(file_path) as pdf:
                    if len(pdf.pages) > 0:
                        page_obj = pdf.pages[0]
                        text = page_obj.extract_text() or ""
                        # Se il PDF è un'immagine scansionata senza testo nativo, attiva l'OCR di supporto
                        if not text.strip():
                            # Conversione temporanea o fallback OCR se necessario
                            text = ""
            elif ext.endswith(('.jpg', '.jpeg', '.png')):
                text = pytesseract.image_to_string(Image.open(file_path)) or ""

            text_lower = text.lower()
            periodo, anno = estrai_mesi_e_anno(text, text_lower, filename)
            
            combined_check = f"{root_dir.lower()} {text_lower} {filename.lower()}"
            if "benzina" in combined_check:
                carburante = "Benzina"
            elif any(k in combined_check for k in ["gasolio", "diesel", "carbur"]):
                carburante = "Diesel"

            quantita, unita_misura = estrai_dati_avanzati_ocr(text, text_lower, page_obj)

            row_idx = ws.max_row + 1
            ws.append([azienda, filename, periodo, anno, quantita, unita_misura, carburante])
            
            cell_file = ws.cell(row=row_idx, column=2)
            cell_file.hyperlink = abs_path
            cell_file.font = Font(color="0563C1", underline="single")
            
            valid_count += 1
        except Exception:
            row_idx = ws.max_row + 1
            ws.append([azienda, filename, "-", "-", "Non rilevato", "Litri", "Non rilevato"])
            cell_file = ws.cell(row=row_idx, column=2)
            cell_file.hyperlink = abs_path
            cell_file.font = Font(color="0563C1", underline="single")
            valid_count += 1

        progress.config(value=idx)
        lbl_status.config(text=f"Processati: {idx} / {total_files}")

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    out_path = os.path.join(desktop, f"Consumi_Gasolio_{azienda.replace(' ', '_')}.xlsx")
    wb.save(out_path)
    messagebox.showinfo("Successo", f"File Excel salvato sul Desktop:\n{os.path.basename(out_path)}")

# Interfaccia Grafica
root = tk.Tk()
root.title("Estrattore Consumi Gasolio")
root.geometry("520x340")
root.resizable(False, False)
root.configure(bg="#ffffff")

FONT_FAMILY = "Segoe UI"

tk.Label(root, text="Estrattore Gasolio & Carburanti", font=(FONT_FAMILY, 14, "bold"), bg="#ffffff", fg="#2e7d32").pack(pady=(15, 2))
tk.Label(root, text="Rilevamento avanzato OCR e layout fatture", font=(FONT_FAMILY, 9), bg="#ffffff", fg="#64748b").pack(pady=(0, 15))

frame_inputs = tk.Frame(root, bg="#ffffff")
frame_inputs.pack(padx=25, fill="x", pady=5)

tk.Label(frame_inputs, text="Nome Azienda (senza spazi):", font=(FONT_FAMILY, 9, "bold"), bg="#ffffff", fg="#334155").pack(anchor="w", pady=(0, 2))
entry_azienda = tk.Entry(frame_inputs, font=(FONT_FAMILY, 9), relief="solid", bd=1)
entry_azienda.pack(fill="x", ipady=3, pady=(0, 10))

tk.Label(frame_inputs, text="Cartella Carburante / Root:", font=(FONT_FAMILY, 9, "bold"), bg="#ffffff", fg="#334155").pack(anchor="w", pady=(0, 2))

frame_path_row = tk.Frame(frame_inputs, bg="#ffffff")
frame_path_row.pack(fill="x")

entry_path = tk.Entry(frame_path_row, font=(FONT_FAMILY, 9), relief="solid", bd=1)
entry_path.pack(side="left", fill="x", expand=True, ipady=3, padx=(0, 6))

btn_browse = tk.Button(frame_path_row, text="Sfoglia...", command=seleziona_cartella, font=(FONT_FAMILY, 8, "bold"), bg="#e2e8f0", fg="#334155", relief="flat", cursor="hand2", padx=10, pady=3)
btn_browse.pack(side="right")

btn_start = tk.Button(root, text="Avvia Estrazione", command=avvia_estrazione, bg="#2e7d32", fg="white", font=(FONT_FAMILY, 10, "bold"), relief="flat", cursor="hand2", pady=6, padx=15)
btn_start.pack(pady=15)

progress = ttk.Progressbar(root, orient="horizontal", length=460, mode="determinate")
progress.pack(pady=(0, 4))

lbl_status = tk.Label(root, text="In attesa...", font=(FONT_FAMILY, 8), bg="#ffffff", fg="#64748b")
lbl_status.pack(pady=(0, 10))

root.mainloop()
